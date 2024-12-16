import os
from openpyxl import load_workbook
from langchain_community.llms import Tongyi
# 设置通义千问API密钥
os.environ["DASHSCOPE_API_KEY"] = "sk-d3361682da654140bde54d50d801ab6c"

llm = Tongyi(model_name='qwen-max')

# Excel文件路径
file_path_feedback = os.path.join(os.path.dirname(__file__), 'data', 'file_path_feedback.xlsx')
wb = load_workbook(file_path_feedback)
sheet = wb.active

for row in sheet.iter_rows(values_only=True, min_row=2):
    feedback = row[0]
    result = llm.invoke(f"""你需要对用户的反馈进行原因分类。
    分类包括：价格过高、售后支持不足、产品使用体验不佳、其他。
    回答格式为：分类结果：xx。
    用户的问题是：{feedback}""")
    print(feedback,result)